import json
import openpyxl

wb = openpyxl.load_workbook('MSFS24_World_Golf_Tour.xlsx', data_only=True)

CATEGORIES = ['GA / Light Jet', 'Regional (ATR/ERJ)', 'Narrowbody (A320/737)', 'Widebody (A340/A350/777/MD-11)']
OK_FIELD = {
    'GA / Light Jet': 'gaOk',
    'Regional (ATR/ERJ)': 'regionalOk',
    'Narrowbody (A320/737)': 'narrowbodyOk',
    'Widebody (A340/A350/777/MD-11)': 'widebodyOk',
}


def to_bool(v):
    if v is None:
        return None
    return str(v).strip().lower() in ('yes', 'ja', 'true', '1', 'y')


# ---------- Tour (courses) ----------
courses = []
ws = wb['Tour']
for row in ws.iter_rows(values_only=True):
    if row[0] is None or not isinstance(row[0], int):
        continue
    (cid, played, region, name, country, hub, closer, arrival, runway,
     ga, regional, narrowbody, widebody, category, rating, ranking, majors) = row
    courses.append({
        'id': cid,
        'played': to_bool(played),
        'region': region,
        'name': name,
        'country': country,
        'hub': hub,
        'closer': closer if closer != '—' else None,
        'arrival': arrival,
        'runway': runway,
        'gaOk': to_bool(ga),
        'regionalOk': to_bool(regional),
        'narrowbodyOk': to_bool(narrowbody),
        'widebodyOk': to_bool(widebody),
        'category': category,
        'rating': rating,
        'ranking': ranking,
        'majors': majors,
    })

# ---------- Flight Legs ----------
legs = []
ws2 = wb['Flight Legs']
for row in ws2.iter_rows(values_only=True):
    if row[0] is None or not isinstance(row[0], int):
        continue
    (lid, frm, frm_name, to, to_name, frm_lat, frm_lon, to_lat, to_lon,
     dist, runway, r_ok, n_ok, w_ok, category, speed, flight_time) = row
    legs.append({
        'id': lid,
        'from': frm,
        'fromName': frm_name,
        'to': to,
        'toName': to_name,
        'fromLat': frm_lat,
        'fromLon': frm_lon,
        'toLat': to_lat,
        'toLon': to_lon,
        'distance': dist,
        'runway': runway,
        'gaOk': True,
        'regionalOk': to_bool(r_ok),
        'narrowbodyOk': to_bool(n_ok),
        'widebodyOk': to_bool(w_ok),
        'category': category,
        'speed': speed,
        'flightTime': flight_time,
    })

# ---------- Settings: fleet + thresholds ----------
ws3 = wb['Settings']
fleet = {}
thresholds = {}
for row in ws3.iter_rows(values_only=True):
    if not row[0] or not isinstance(row[0], str):
        continue
    if row[0] in CATEGORIES and len(row) >= 3 and row[2] is not None:
        fleet[row[0]] = {'planes': row[1], 'speed': row[2]}
    if row[0].endswith('minimum'):
        cat = row[0][:-len(' minimum')].strip()
        if cat in CATEGORIES:
            thresholds[cat] = {'min': row[1], 'note': row[2] if len(row) > 2 else None}

# ---------- Airport coords from legs ----------
airports = {}
for l in legs:
    airports.setdefault(l['from'], (l['fromLat'], l['fromLon']))
    airports.setdefault(l['to'], (l['toLat'], l['toLon']))

missing = []
for c in courses:
    coord = airports.get(c['arrival'])
    if coord is None:
        missing.append(c['arrival'])
        coord = (None, None)
    c['lat'], c['lon'] = coord

regions = []
for c in courses:
    if c['region'] not in regions:
        regions.append(c['region'])

# ---------- Dashboard stats ----------
wsd = wb['Dashboard']
rows = list(wsd.iter_rows(values_only=True))


def find_value(label):
    for ri, r in enumerate(rows):
        for ci, cell in enumerate(r):
            if cell is not None and str(cell).strip().lower() == label.lower():
                if ri + 1 < len(rows) and ci < len(rows[ri + 1]):
                    return rows[ri + 1][ci]
    return None


stats = {
    'played': find_value('Courses Played'),
    'completion': find_value('Completion %'),
    'totalLegs': find_value('Total Legs Flown'),
    'distanceNm': find_value('Total Distance (nm)'),
    'flightTimeH': find_value('Total Flight Time'),
    'distanceKm': find_value('Total Distance (km)'),
}

# Legs by recommended category (counted from the legs data)
legs_by_cat = {c: 0 for c in CATEGORIES}
for l in legs:
    if l['category'] in legs_by_cat:
        legs_by_cat[l['category']] += 1

data = {
    'categories': CATEGORIES,
    'okField': OK_FIELD,
    'fleet': fleet,
    'thresholds': thresholds,
    'courses': courses,
    'legs': legs,
    'airports': airports,
    'regions': regions,
    'stats': stats,
    'legsByCategory': legs_by_cat,
    'missing': missing,
}

with open('data.js', 'w', encoding='utf-8') as f:
    f.write('// Auto-generated from MSFS24_World_Golf_Tour.xlsx\n')
    f.write('const TOUR_DATA = ' + json.dumps(data, ensure_ascii=False, indent=1) + ';\n')

print('courses:', len(courses))
print('legs:', len(legs))
print('airports:', len(airports))
print('fleet:', list(fleet.keys()))
print('thresholds:', {k: v['min'] for k, v in thresholds.items()})
print('legsByCategory:', legs_by_cat)
print('missing coords:', missing)
